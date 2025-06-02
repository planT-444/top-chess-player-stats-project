import openpyxl
from pathlib import Path
from collections import defaultdict
import json

filepath = Path(__file__).parent / 'ratings.xlsx'
sheet = openpyxl.load_workbook(filepath).active
if sheet is None:
    raise ValueError("No active sheet in workbook")

rating_sums = defaultdict(int)

for cindex in (1, 4, 7):
    for rindex in range(2, 42):
        name = sheet.cell(row = rindex, column = cindex).value
        rating = sheet.cell(row = rindex, column = cindex + 1).value
        if isinstance(rating, (float, int)):
            rating_sums[name] += int(rating)


sorted_ratings = sorted(rating_sums.items(), key = lambda x: x[1], reverse = True)[:20]
for player, rating_sum in sorted_ratings:
    print(f"{player}: {rating_sum}")
with open("player_ratings.json", 'w') as f:
    json.dump(dict(sorted_ratings), f, indent = 4)