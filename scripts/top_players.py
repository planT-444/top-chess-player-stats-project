import openpyxl
from pathlib import Path
from collections import defaultdict
import json
root_dir = Path(__file__).parent.parent
filepath = root_dir / "input-files" / "ratings.xlsx"
output_filepath = root_dir / "output-files" / "player_ratings.json"
sheet = openpyxl.load_workbook(filepath).active
if sheet is None:
    raise ValueError("No active sheet in workbook")

rating_sums = defaultdict(int)
classical = {}
rapid = {}
blitz = {}

for cindex in (1, 4, 7):
    for rindex in range(2, 42):
        name = sheet.cell(row = rindex, column = cindex).value
        rating = sheet.cell(row = rindex, column = cindex + 1).value
        if isinstance(rating, (float, int)):
            rating_sums[name] += int(rating)
            match cindex:
                case 1:
                    blitz[name] = int(rating)
                case 4:
                    rapid[name] = int(rating)
                case 7:
                    classical[name] = int(rating)

sorted_ratings = sorted(rating_sums.items(), key = lambda x: x[1], reverse = True)[:21]
for player, rating_sum in sorted_ratings:
    print(f"{player}: {rating_sum}")
print("| yeet |")
print("|------|")
for player, rating_sum in sorted_ratings:
    print(f'| {blitz[player]} |')
# with open(output_filepath, 'w') as f:
#     json.dump(dict(sorted_ratings), f, indent = 4)